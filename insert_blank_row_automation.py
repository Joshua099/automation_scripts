"""
Created by Joshua Pantoja.

Quick script to automate table formatting for a report requested by manager.

Make sure to sort customerID in ascending order first and convert table to range.
"""

import openpyxl as op

wb = op.load_workbook('WOES_Albuquerque_contracted_pricing_report_v1.xlsx')

# column 1 = CustomerID
col = 1

# With column 1 as the index, iterates through specified amount of rows. Checks if a cell's CustomerID =/= the cell
# above the active cell. If it is not equal, then add a new blank row above it. Afterwords, it's easy to filter blank
# rows and format them with the neutral format color for the report.


# Alter max range value to account for # of rows in spreadsheet.
for row in range(2, 1000):  # ## for sheet [1]
    if not wb.worksheets[0].cell(row, col).value == wb.worksheets[0].cell(row-1, col).value:
        if not wb.worksheets[0].cell(row-1, col).value is None:
            wb.worksheets[0].insert_rows(row)

wb.save('output.xlsx')

wb.close()

""" 
Previous attempts. Above code works.
"""

# ws.worksheets[0].iter_rows(min_row=2, max_row=10000, min_col=col, max_col=col):

# # Need a way of telling excel to insert a new row ONLY AFTER a group of the same values (one customer)
# list = []
# for row in ws.iter_rows(min_row=2,max_row=4543,min_col=1,max_col=1):
#     for cell in row:
#         list.append(cell.value)

# unique_values = []
#
# # Now have a list of all the unique customers
# for x in list:
#     if x not in unique_values:
#         unique_values.append(x)
#
# print(unique_values)
# print(len(unique_values))

# Need a way of inserting a new row, where the specifying index is only after a group of one value

# for row in ws.iter_rows(min_row=2,max_row=4543,min_col=1,max_col=1):
#     for cell in row:
#         for x in unique_values:
#             if cell.value == x:
#                 print(cell)



